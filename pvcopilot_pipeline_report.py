"""
pvcopilot_pipeline_report.py
============================

End-to-end automation harness for the PVcopilot tool.

For every dataset in a folder (default ./test_datasets/), this drives the SAME
pipeline the website runs -- Analyze -> Apply Filters -> Calculate Degradation --
using the exact functions the app's callbacks call, and records, per dataset:

    GOOD     full pipeline ran and produced a finite degradation rate with no
             data-quality caveats (purely informational notes -- e.g. "multiple
             columns matched, picked one" -- do NOT demote a dataset)
    WARNING  ran, but with caveats (no irradiance, no temperature, YoY disabled,
             unreliable rate, sparse/missing/near-constant columns, AC fallback)
    BLOCKED  tool correctly refused (under 1 year of data -- expected behavior)
    ERROR    the tool failed -- file unreadable, column mapping failed, a filter
             or the degradation method crashed, timed out, or returned no rate

Each run writes a timestamped, machine-readable report AND a human-readable PDF
into ./reports/ , plus a "latest" copy that always points at the most recent run:

Files are numbered per run (1, 2, 3, ...) so the HIGHEST number is the latest run:
    reports/pvcopilot_pipeline_report_<N>.json   (full detail)
    reports/pvcopilot_pipeline_report_<N>.csv    (one row per dataset, status/stage/reason)
    reports/pvcopilot_results_<N>.csv            (focused results sheet)

The focused results sheet has columns:
    dataset | total_duration_years | mean_power_w | filtering_result |
    degradation_rate_pct_per_year | rate_reliable | unreliable_because |
    missing_maps | multiple_matches | statistical_trend_method |
    detected_or_created_columns | why_missing
(unreliable_because spells out exactly which reliability checks failed when
rate_reliable is "no" — empty when the rate is reliable. Missing irradiance
is NOT counted against rate_reliable in these reports — it appears as a
"not weather-normalized" note instead; the website's verdict stays strict.)
(multiple_matches lists, per variable, every column that matched when there
was more than one — with the chosen one marked "(used)".)
(mean_power_w is in watts; units are inferred from the power column name.)

Usage (always with the pvtools env):
    conda activate pvtools
    python pvcopilot_pipeline_report.py                  # scans ./test_datasets/
    python pvcopilot_pipeline_report.py path/to/folder   # scan a different folder
    python pvcopilot_pipeline_report.py --include-examples   # also test data/*.parquet
    python pvcopilot_pipeline_report.py --method LR      # force a degradation method

The classification mirrors the app's own gating:
  * < 1 year of data           -> BLOCKED (degradation not attempted)
  * 1 <= years < 2             -> YoY disabled, falls back to Linear Regression
  * >= 2 years                 -> YoY (the app's default method)
"""

import os
import re
import sys
import csv
import json
import base64
import fnmatch
import argparse
import textwrap
import traceback
from contextlib import redirect_stdout
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout

# --- make the project root importable, then load API keys like index.py does ---
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

# All imports below are third-party / app modules that only exist in the
# `pvtools` conda env. If you launch this with the system Python (e.g.
# `/usr/bin/python3 ...`) they won't be found -- so turn the raw
# ModuleNotFoundError into a clear instruction instead.
try:
    from dotenv import load_dotenv

    import numpy as np
    import pandas as pd
    from tqdm import tqdm

    # The SAME functions PVcopilot's callbacks use.
    from page_supporting_files.analysis_utils_2606 import (
        parse_contents, normalize, low_irra_power_filter, low_power_filter,
        aggregate_daily, compute_yoy, compute_lr, compute_hw, compute_arima,
        compute_csd, rate_is_plausible, degradation_reliability, MIN_TREND_POINTS,
        compute_yoy_energy,
    )
    from page_supporting_files.pvcopilot_filter_functions import (
        basic_value_filter, clear_sky_filter, identify_outliers_iqr,
        stale_data_filter, detect_and_fix_time_shifts,
    )
except ModuleNotFoundError as e:
    sys.stderr.write(
        f"\nERROR: missing dependency '{e.name}'.\n"
        f"This script must run inside the 'pvtools' conda env, not the system\n"
        f"Python ({sys.executable}).\n\n"
        f"Fix:\n"
        f"    conda activate pvtools\n"
        f"    python pvcopilot_pipeline_report.py\n\n"
        f"Or call the env's Python directly:\n"
        f"    /opt/anaconda3/envs/pvtools/bin/python pvcopilot_pipeline_report.py\n\n"
    )
    sys.exit(1)

load_dotenv(override=True)  # load API keys like index.py does

SUPPORTED = (".csv", ".xlsx", ".xls", ".parquet")
DEFAULT_DATASET_DIR = os.path.join(PROJECT_ROOT, "test_datasets")
EXAMPLE_DIR = os.path.join(PROJECT_ROOT, "data")
REPORT_DIR = os.path.join(PROJECT_ROOT, "reports")

# Mirror the app: cap each heavy stage so a pathological dataset can't hang the run.
STEP_TIMEOUT_S = 10
# The analyze stage makes an LLM call (bounded client-side in analysis_utils),
# which can spike or retry once on a busy gateway, so give it more headroom than
# the pure-pandas filter/degradation stages.
ANALYZE_TIMEOUT_S = 25
# Minimum rows surviving filtering to even attempt a degradation fit. Below this
# the estimate is meaningless (and the methods can crash), so we refuse cleanly.
MIN_FILTERED_ROWS = 20
# Default filters match the app's `filter-options` default value.
DEFAULT_FILTERS = ["timezone", "low-irra-power", "outlier", "clearsky", "stale"]

STATUS_COLOR = {
    "GOOD":    "#16a34a",  # green
    "WARNING": "#d97706",  # amber
    "BLOCKED": "#2563eb",  # blue (expected refusal, not a failure)
    "ERROR":   "#dc2626",  # red
}
STATUS_ORDER = ["GOOD", "WARNING", "BLOCKED", "ERROR"]

# Notes matching these prefixes describe a SUCCESSFUL, unambiguous decision --
# the mapper narrating its work (picked one of several matching columns,
# derived DC power as V x I, found the time column by value). They stay in the
# JSON/notes for transparency but are NOT data-quality caveats, so they do not
# demote a dataset from GOOD to WARNING. Anything not listed here counts as a
# caveat by default (fail-safe: an unknown note still demotes).
INFO_NOTE_PREFIXES = (
    "Multiple columns matched",
    "DC Power computed as Voltage × Current",
    "DC Voltage computed as Power / Current",
    "DC Current computed as Power / Voltage",
    "Auto-selected irradiance",
    "Time column detected from values",
)


def _caveat_notes(notes):
    """The subset of notes that are genuine data-quality caveats."""
    return [n for n in notes if not n.startswith(INFO_NOTE_PREFIXES)]

_POOL = ThreadPoolExecutor(max_workers=4)


def _with_timeout(fn, *args, timeout=STEP_TIMEOUT_S, **kwargs):
    return _POOL.submit(fn, *args, **kwargs).result(timeout=timeout)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _extract_text(component):
    """Pull readable text out of a Dash component / string / list."""
    if component is None:
        return ""
    if isinstance(component, str):
        return component
    if isinstance(component, (list, tuple)):
        return " ".join(_extract_text(c) for c in component).strip()
    children = getattr(component, "children", None)
    if children is not None:
        return _extract_text(children)
    return ""


def _build_contents(path):
    """Mirror the browser upload: a base64 data-URL string."""
    with open(path, "rb") as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()
    return f"data:application/octet-stream;base64,{b64}"


def _power_unit_scale(col_name):
    """Multiplier to convert a power column to WATTS, inferred from its name
    (units aren't stored in the data): kW -> 1000, MW -> 1e6, else assume watts.
    Computed V x I power ('computed_dc_power') is already in watts."""
    name = (col_name or "").lower()
    if "mw" in name:
        return 1_000_000.0
    if "kw" in name and "kwh" not in name:   # kWh would be energy, not power
        return 1000.0
    return 1.0


def _duration_years(df):
    """Span of the data in years, measured from the time index.

    Only trusts a genuine DatetimeIndex (which parse_contents sets from the
    identified Time column). Returns None when time isn't established -- e.g. no
    Time column was identified -- so we don't mistake a plain integer index for
    a near-zero (1970 epoch) span and block it for the wrong reason.
    """
    idx = df.index
    if not isinstance(idx, pd.DatetimeIndex) or len(idx) == 0:
        return None
    return (idx.max() - idx.min()).days / 365.25


def _run_filters(df, mapping):
    """Replicate the app's run_filter compute path with default parameters.
    Returns the filtered, normalized dataframe (rows kept)."""
    power_key = mapping.get("DC Power")
    irra_key = mapping.get("Irradiance")
    has_irr = bool(irra_key) and irra_key in df.columns

    bv_normal, _ = basic_value_filter(df, mapping)
    df = df.loc[bv_normal].copy()

    clearsky_mask = pd.Series(True, index=df.index)
    if "clearsky" in DEFAULT_FILTERS and has_irr:
        normal_idx, _ = clear_sky_filter(df, irra_key,
                                         smoothness_threshold=0.3,
                                         energy_threshold=0.5)
        clearsky_mask = df.index.isin(normal_idx)

    # NEW-FILTERS: stale mask on the raw signals, before normalization and the
    # timezone relabel (same order as the app's run_filter).
    stale_mask = pd.Series(True, index=df.index)
    if "stale" in DEFAULT_FILTERS:
        try:
            normal_idx, _ = stale_data_filter(df, mapping, window=6)
            stale_mask = df.index.isin(normal_idx)
        except Exception:
            pass  # stale detection is best-effort in the report too

    df_f = normalize(df, mapping, gamma=-0.004)
    current_mask = pd.Series(
        pd.Series(clearsky_mask).values & pd.Series(stale_mask).values,
        index=df_f.index)

    # TIME-SHIFT: data-driven correction, exactly as the app's run_filter does it.
    if "timezone" in DEFAULT_FILTERS:
        try:
            df_f.index = pd.to_datetime(df_f.index)
            df_f, _tz_msg = detect_and_fix_time_shifts(df_f, mapping.get("DC Power"))
            current_mask.index = df_f.index
        except Exception:
            pass  # non-datetime index; app tolerates this too

    if "low-irra-power" in DEFAULT_FILTERS and has_irr:
        normal_idx, _ = low_irra_power_filter(
            df_f, mapping, irr_thresh=300, power_ratio=0.02,
            norm_lower=0.01, norm_upper_pct=99)
        current_mask &= df_f.index.isin(normal_idx)
    elif not has_irr:
        # No irradiance: drop night / low-output points by power alone.
        normal_idx, _ = low_power_filter(df_f, mapping)
        current_mask &= df_f.index.isin(normal_idx)

    if "outlier" in DEFAULT_FILTERS:
        # Compute IQR fences on the points that survived the prior filters, not
        # the full frame -- otherwise night/low-output values dominate the
        # distribution and real daytime production gets flagged as outliers
        # (which, intersected with low_power_filter, can wipe out every row).
        kept = df_f.loc[df_f.index[current_mask]]
        normal_idx, _ = identify_outliers_iqr(kept, "norm", iqr_multiplier=1.5)
        current_mask &= df_f.index.isin(normal_idx)

    return df_f.loc[df_f.index[current_mask]]


def _compute_one(daily, method):
    if method == "YOY":
        return compute_yoy(daily)[0]
    if method == "LR":
        return compute_lr(daily)[0]
    if method == "HW":
        return compute_hw(daily, period=12)[0]
    if method == "ARIMA":
        return compute_arima(daily, p=1, d=1, q=0, seasonal_period=12)[0]
    if method == "CSD":
        return compute_csd(daily, period=12)[0]
    raise ValueError(f"Unknown method: {method}")


def _run_degradation(df_filtered, mapping, method):
    """aggregate_daily + the chosen method, with an LR fallback. Returns
    (rd, method_used). YoY/HW/ARIMA/CSD need regular, dense, year-spanning data;
    on sparse/irregular series they return NaN. Rather than failing a perfectly
    good dataset, fall back to Linear Regression (which fits a trend to any set
    of points). Returns method_used so the caller can note the fallback."""
    irra_key = mapping.get("Irradiance")
    if not irra_key or irra_key not in df_filtered.columns:
        irra_key = None
    daily = aggregate_daily(df_filtered, irra_key)
    rd = _compute_one(daily, method)
    if (rd is None or not np.isfinite(rd)) and method != "LR":
        rd_lr = _compute_one(daily, "LR")
        if rd_lr is not None and np.isfinite(rd_lr):
            return rd_lr, "LR"
    return rd, method


# ---------------------------------------------------------------------------
# Per-dataset pipeline
# ---------------------------------------------------------------------------
def run_one(path, forced_method=None):
    """Drive Analyze -> Filter -> Degradation for one dataset.

    Returns a dict capturing the outcome, the stage reached, the reason, and
    useful metadata.
    """
    fname = os.path.basename(path)
    rec = {
        "file": fname, "status": "ERROR", "stage": "load", "reason": "",
        "rows_raw": None, "rows_filtered": None, "filtering_result": None,
        "duration_years": None, "mean_power_w": None, "mapped": {}, "method": None,
        "rd_percent_per_year": None, "rate_reliable": None, "notes": [],
        # {role: [other valid columns]} when several columns matched one
        # variable — surfaced as the multiple_matches column in the results CSV.
        "alternatives": {},
    }

    # ---- Stage 1: ANALYZE (load + LLM column identification) ----
    try:
        contents = _build_contents(path)
        df, summary_table, mapping, _code, mapping_notes = _with_timeout(
            parse_contents, contents, fname, timeout=ANALYZE_TIMEOUT_S)
    except FutureTimeout:
        rec.update(stage="analyze", reason=f"Analyze exceeded {STEP_TIMEOUT_S}s timeout.")
        return rec
    except Exception as e:
        rec.update(stage="analyze",
                   reason=f"{type(e).__name__}: {e}\n" + traceback.format_exc(limit=2))
        return rec
    # AC-fallback / computed-power / time-by-value warnings from parse_contents.
    # All are recorded in the JSON; only the ones that are genuine data-quality
    # caveats (see INFO_NOTE_PREFIXES) demote the row to WARNING.
    if mapping_notes:
        rec["notes"].extend(mapping_notes)

    # Column ambiguity is no longer a mapping_note (it's a UI-only inline hint),
    # so read it from df.attrs to keep flagging those rows in the batch report.
    alts = getattr(df, "attrs", {}).get("mapping_alternatives") if df is not None else None
    if alts:
        rec["alternatives"] = alts
        for role, others in alts.items():
            rec["notes"].append(
                f"Multiple columns matched {role}; also valid: {', '.join(others)}.")

    if df is None:
        # A genuine read/parse failure is an ERROR (the tool broke on the file).
        rec.update(stage="analyze",
                   reason=_extract_text(summary_table).strip() or "File could not be read/parsed.")
        return rec
    if not mapping:
        # No PV variables could be identified -> not enough information to solve,
        # which is a clean refusal (BLOCKED), not a tool failure.
        rec.update(status="BLOCKED", stage="analyze",
                   reason="Not enough information: no usable PV variables could be "
                          "identified in this dataset.")
        return rec

    rec["mapped"] = dict(mapping)
    rec["rows_raw"] = int(len(df))

    # Mean DC power across ALL parsed rows (pre-filter), normalized to WATTS.
    # Units aren't in the data, so we infer them from the power column's name
    # (kW/MW -> scaled; otherwise watts). Computed V x I power is already watts.
    # Captured here so even BLOCKED datasets report it.
    _pk = mapping.get("DC Power")
    if _pk and _pk in df.columns:
        try:
            _p = pd.to_numeric(df[_pk], errors="coerce")
            _mean = float(_p.mean())
            rec["mean_power_w"] = round(_mean * _power_unit_scale(_pk), 1) if np.isfinite(_mean) else None
            # Near-constant power (tiny coefficient of variation) means the column
            # is daily-aggregated / a non-instantaneous channel: there's no
            # night/cloud variation to filter, so retention will look very high
            # AND the degradation signal is weak. Flag it so a 99%-retained row
            # isn't mistaken for "99% clear-sky-quality data".
            _cov = float(_p.std() / _mean) if _mean else 0.0
            if 0 < _cov < 0.1:
                rec["notes"].append(
                    f"Power column '{_pk}' is near-constant (variation {_cov:.1%}) — "
                    "looks daily-aggregated, not instantaneous. Little to filter "
                    "(retention will be high) and the degradation signal is weak.")
        except Exception:
            rec["mean_power_w"] = None

    # ---- Stage 2 (FIRST GATE): DURATION -------------------------------------
    # Block purely on HOW LONG the data spans, before requiring any specific
    # column. Data under a year can never yield a meaningful degradation rate,
    # so "too short" takes precedence over "missing columns": a single-day file
    # reports BLOCKED (expected refusal), not ERROR. This needs only the time
    # index, which parse_contents sets from the identified Time column.
    dur = _duration_years(df)
    rec["duration_years"] = round(dur, 3) if dur is not None else None
    if dur is not None and dur < 1.0:
        months = int(round(dur * 12))
        rec.update(status="BLOCKED", stage="duration",
                   reason=f"Only ~{months} months (<1 year). Degradation correctly not attempted.")
        return rec

    # ---- Required columns (only reached by data that is long enough) --------
    # No usable Time/DC Power column means there isn't enough information to
    # solve the problem -> BLOCKED (clean refusal), not ERROR (a tool failure).
    missing_required = [v for v in ("DC Power", "Time") if v not in mapping]
    if missing_required:
        rec.update(status="BLOCKED", stage="analyze",
                   reason="Not enough information: no usable "
                          + " and ".join(missing_required)
                          + " column found, so degradation can't be computed.")
        return rec

    # Data-quality notes. (The "power computed as V x I" note already comes from
    # parse_contents via mapping_notes, so we don't add a duplicate here.)
    has_irr = bool(mapping.get("Irradiance")) and mapping["Irradiance"] in df.columns
    has_temp = bool(mapping.get("Module temperature")) and mapping["Module temperature"] in df.columns
    if not has_irr:
        rec["notes"].append("No irradiance -> not weather-normalized (less reliable).")
    elif not has_temp:
        rec["notes"].append("No module temperature -> no temperature correction.")

    # Method follows the app's gating: YoY needs >= 2 years; else LR.
    if forced_method:
        method = forced_method
    elif dur is not None and dur < 2.0:
        method = "LR"
        rec["notes"].append("Under 2 years -> YoY disabled, using Linear Regression.")
    else:
        method = "YOY"
    rec["method"] = method

    # ---- Stage 3: FILTER ----
    try:
        df_filtered = _with_timeout(_run_filters, df, mapping)
    except FutureTimeout:
        rec.update(stage="filter", reason=f"Filtering exceeded {STEP_TIMEOUT_S}s timeout.")
        return rec
    except Exception as e:
        rec.update(stage="filter",
                   reason=f"{type(e).__name__}: {e}\n" + traceback.format_exc(limit=2))
        return rec

    rec["rows_filtered"] = int(len(df_filtered))
    # Filtering result: share of the parsed rows that survived the filter chain
    # (basic-value + clear-sky + low-irradiance + IQR), like the app's "% retained".
    if rec["rows_raw"]:
        pct = 100.0 * rec["rows_filtered"] / rec["rows_raw"]
        rec["filtering_result"] = (f"{pct:.1f}% retained "
                                   f"({rec['rows_filtered']}/{rec['rows_raw']} rows)")
    if len(df_filtered) < MIN_FILTERED_ROWS:
        rec.update(status="BLOCKED", stage="filter",
                   reason=f"Only {len(df_filtered)} row(s) survived filtering "
                          f"(need >= {MIN_FILTERED_ROWS}) -- not enough clean data to "
                          "estimate degradation. Correctly refused rather than guessing.")
        return rec

    # Count AGGREGATED DAILY points -- drives the reliability flag below. A trend
    # from a few daily points is still reported, but flagged as unreliable (with
    # the reason). Only a truly uncomputable case (<2 points) is refused.
    _irr_key = mapping.get("Irradiance")
    _irr_key = _irr_key if _irr_key and _irr_key in df_filtered.columns else None
    _daily = aggregate_daily(df_filtered, _irr_key)
    _n_daily = int(_daily.dropna().shape[0]) if _daily is not None else 0
    if _n_daily < 2:
        rec.update(status="BLOCKED", stage="degradation",
                   reason=f"Only {_n_daily} daily data point(s) after aggregation -- "
                          "a trend cannot be computed at all.")
        return rec

    # ---- Stage 4: DEGRADATION ----
    # Preferred: the energy-ratio YoY (sums each day's measured vs expected
    # energy BEFORE trending — steadier, uses more of the data). Only the rate
    # is reported; its internal uncertainty band feeds the reliability flag.
    # Falls back to the classic path (point-ratio daily aggregate, LR fallback)
    # when the energy method can't run (no irradiance / too few paired days).
    _ci_width = None
    rd, method_used = None, method
    if method == "YOY" and _irr_key:
        try:
            _rd_e, _fig_e, _ciw, _ndays = _with_timeout(
                compute_yoy_energy, df, mapping)
            if _rd_e is not None and np.isfinite(_rd_e):
                rd, method_used = float(_rd_e), "YOY"
                rec["method"] = "YOY (energy-ratio)"
                _ci_width, _n_daily = _ciw, _ndays
        except Exception as e:
            rec["notes"].append(
                f"Energy-ratio YoY errored ({type(e).__name__}: {e}) — "
                "fell back to the classic path.")

    if rd is None:
        try:
            rd, method_used = _with_timeout(_run_degradation, df_filtered, mapping, method)
        except FutureTimeout:
            rec.update(stage="degradation", reason=f"Degradation exceeded {STEP_TIMEOUT_S}s timeout.")
            return rec
        except Exception as e:
            rec.update(stage="degradation",
                       reason=f"{type(e).__name__}: {e}\n" + traceback.format_exc(limit=2))
            return rec

    if method_used != method:
        rec["notes"].append(
            f"{method} produced no rate (data too sparse/irregular for it); "
            f"fell back to Linear Regression.")
        method = method_used
        rec["method"] = method_used

    if rd is None or not np.isfinite(rd):
        rec.update(status="ERROR", stage="degradation",
                   reason=f"{method} returned no finite degradation rate (got {rd}).")
        return rec

    rec["rd_percent_per_year"] = round(float(rd), 4)
    # Report the rate but flag whether it's trustworthy, with the specific
    # reason(s) -- too few daily points, no irradiance, short span, an
    # implausible value, or (energy path) too-scattered year-over-year changes.
    reliable, reasons = degradation_reliability(
        rd, n_points=_n_daily, has_irradiance=bool(_irr_key),
        duration_years=rec.get("duration_years"), ci_width=_ci_width)
    # REPORT-ONLY POLICY: missing irradiance alone does not disqualify a rate
    # here — it is already surfaced as a "not weather-normalized" note above.
    # Only hard failures (implausible value, scattered YoY, too few points,
    # short span) count against rate_reliable in these reports. The website's
    # degradation_reliability verdict stays strict and is NOT affected.
    reasons = [x for x in reasons if "irradiance" not in x]
    # REPORT-ONLY POLICY: the scattered-YoY check only counts when the
    # year-over-year disagreement exceeds ±2 %/yr (ci_width > 4). The website
    # keeps the stricter ±1.5 (MAX_CI_WIDTH = 3) verdict.
    REPORT_MAX_CI_WIDTH = 4.0
    if _ci_width is not None and np.isfinite(_ci_width) and _ci_width <= REPORT_MAX_CI_WIDTH:
        reasons = [x for x in reasons if "disagree" not in x]
    reliable = len(reasons) == 0
    rec["rate_reliable"] = reliable
    rec["reliability_reasons"] = reasons   # -> unreliable_because in results CSV
    if not reliable:
        rec["notes"].append(
            f"Rate {rd:+.2f}%/yr is UNRELIABLE because " + "; ".join(reasons) + ".")

    rec["stage"] = "done"
    # Only genuine data-quality caveats demote to WARNING. Informational
    # mapper-narration notes ("multiple columns matched", "power computed as
    # V x I") used to make GOOD unreachable -- every real SCADA export has at
    # least one -- so a binary any-note rule marked ALL datasets WARNING.
    caveats = _caveat_notes(rec["notes"])
    rec["status"] = "WARNING" if caveats else "GOOD"
    rec["reason"] = (f"{method} degradation rate = {rd:.3f} %/yr"
                     + (" (with caveats)" if caveats else ""))
    return rec


# ---------------------------------------------------------------------------
# Report writers
# ---------------------------------------------------------------------------
def _counts(results):
    c = {k: 0 for k in STATUS_ORDER}
    for r in results:
        c[r["status"]] = c.get(r["status"], 0) + 1
    return c


def write_json(results, counts, meta, path):
    with open(path, "w") as f:
        json.dump({"meta": meta, "summary": counts, "results": results}, f, indent=2)


def write_csv(results, path):
    cols = ["file", "status", "stage", "duration_years", "method",
            "rd_percent_per_year", "rows_raw", "rows_filtered", "reason"]
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in results:
            row = dict(r)
            row["reason"] = " ".join(str(row.get("reason", "")).split())  # flatten newlines
            w.writerow(row)


# The variables PVcopilot tries to detect in every dataset.
EXPECTED_ROLES = ["Time", "DC Power", "DC Voltage", "DC Current",
                  "Irradiance", "Module temperature"]


def _format_detected_columns(mapped):
    """Render the variables the tool detected (or created) from the dataset's
    columns, e.g. 'DC Power=power | Irradiance=irr | Time=measured_on'. A value
    of 'computed_dc_power' means power was computed as Voltage x Current."""
    if not mapped:
        return ""
    # Stable, readable order; any extra roles appended after the known ones.
    keys = ([k for k in EXPECTED_ROLES if k in mapped]
            + [k for k in mapped if k not in EXPECTED_ROLES])
    return " | ".join(f"{k}={mapped[k]}" for k in keys)


def _missing_maps(mapped):
    """Which of the expected variables the tool could NOT map for this dataset,
    e.g. 'Irradiance, Module temperature'. 'none' when all were found."""
    mapped = mapped or {}
    missing = [role for role in EXPECTED_ROLES if role not in mapped]
    return ", ".join(missing) if missing else "none"


def _why_missing(r):
    """Explain why the filtering result and/or degradation rate is blank for a
    dataset, using the stage the pipeline stopped at and its reason. Empty when
    both values are present (nothing missing)."""
    has_filter = bool(r.get("filtering_result"))
    has_rd = r.get("rd_percent_per_year") is not None
    if has_filter and has_rd:
        return ""
    missing = []
    if not has_filter:
        missing.append("filtering result")
    if not has_rd:
        missing.append("degradation rate")
    reason = " ".join(str(r.get("reason", "")).split())  # flatten newlines
    stage = r.get("stage")
    return f"No {' and '.join(missing)} (stopped at '{stage}'): {reason}"


def _format_multi_matches(rec):
    """When several columns matched one variable, list them per role with the
    one the tool chose marked '(used)'. 'none' when every role had a single
    match. Example:
        DC Power: dc_power__2841 (used), inv1_dc_power__2811 | Irradiance: ...
    """
    alts = rec.get("alternatives") or {}
    if not alts:
        return "none"
    mapped = rec.get("mapped") or {}
    parts = []
    for role, others in alts.items():
        chosen = mapped.get(role)
        cols = ([f"{chosen} (used)"] if chosen else []) + list(others)
        parts.append(f"{role}: {', '.join(cols)}")
    return " | ".join(parts)


RESULTS_COLS = ["dataset", "total_duration_years", "mean_power_w",
                "filtering_result", "degradation_rate_pct_per_year",
                "rate_reliable", "unreliable_because", "missing_maps",
                "multiple_matches", "statistical_trend_method",
                "detected_or_created_columns", "why_missing"]


def _results_row(r):
    """One dataset's results-sheet row (shared by the CSV and XLSX writers)."""
    return {
        "dataset": r["file"],
        "total_duration_years":
            "" if r.get("duration_years") is None else r["duration_years"],
        "mean_power_w":
            "" if r.get("mean_power_w") is None else r["mean_power_w"],
        "filtering_result": r.get("filtering_result") or "",
        "degradation_rate_pct_per_year":
            "" if r.get("rd_percent_per_year") is None else r["rd_percent_per_year"],
        "rate_reliable":
            "" if r.get("rate_reliable") is None
            else ("yes" if r["rate_reliable"] else "no"),
        "unreliable_because": "; ".join(r.get("reliability_reasons") or []),
        "missing_maps": _missing_maps(r.get("mapped")),
        "multiple_matches": _format_multi_matches(r),
        "statistical_trend_method": r.get("method") or "",
        "detected_or_created_columns": _format_detected_columns(r.get("mapped")),
        "why_missing": _why_missing(r),
    }


def write_results_csv(results, path):
    """Focused results sheet: one row per dataset with the tool's actual outputs
    -- the filtering result and degradation rate (when they exist), the columns
    the tool detected/created, and (when either output is missing) why."""
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=RESULTS_COLS)
        w.writeheader()
        for r in results:
            w.writerow(_results_row(r))


def write_results_xlsx(results, path):
    """The same results sheet as a READABLE Excel file: every cell wraps its
    text, columns have sensible widths, the header row is bold and frozen, and
    rows whose rate is flagged unreliable get a light amber tint. (A plain CSV
    can't carry any formatting, so this is the copy meant for human eyes.)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return False   # openpyxl not installed -- CSV/JSON still written

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    widths = {  # tuned per column: long free-text columns get wide + wrap
        "dataset": 30, "total_duration_years": 10, "mean_power_w": 12,
        "filtering_result": 24, "degradation_rate_pct_per_year": 12,
        "rate_reliable": 9, "unreliable_because": 52, "missing_maps": 26,
        "multiple_matches": 46, "statistical_trend_method": 16,
        "detected_or_created_columns": 46, "why_missing": 46,
    }
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.append(RESULTS_COLS)
    for i, c in enumerate(RESULTS_COLS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True)
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 18)
    ws.freeze_panes = "A2"

    amber = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
    for r in results:
        row = _results_row(r)
        ws.append([row[c] for c in RESULTS_COLS])
        for i in range(1, len(RESULTS_COLS) + 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.alignment = wrap
            if row["rate_reliable"] == "no":
                cell.fill = amber
        # openpyxl does NOT auto-fit wrapped rows (and neither do many
        # viewers), so without an explicit height the wrapped lines hide
        # behind a one-line-tall row. Estimate how many lines the longest
        # cell wraps into (~1.9 characters per unit of column width) and
        # size the row to show ALL of them.
        max_lines = 1
        for c in RESULTS_COLS:
            text = str(row[c])
            if not text:
                continue
            chars_per_line = max(int(widths.get(c, 18) * 1.9), 8)
            max_lines = max(max_lines, -(-len(text) // chars_per_line))
        ws.row_dimensions[ws.max_row].height = min(15 * max_lines + 4, 220)
    wb.save(path)
    return True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def _next_run_number():
    """Next sequential run number: 1 + the highest already in reports/.
    Numbered filenames mean the highest number is always the latest run."""
    n = 0
    if os.path.isdir(REPORT_DIR):
        for f in os.listdir(REPORT_DIR):
            m = re.match(r"pvcopilot_(?:results|pipeline_report)_(\d+)\.", f)
            if m:
                n = max(n, int(m.group(1)))
    return n + 1


def collect_files(dataset_dir, include_examples, match=None):
    """Collect supported dataset files, optionally keeping only filenames that
    match a glob pattern (e.g. 'my_*' to test just the combined my_* datasets)."""
    def keep(f):
        if not f.lower().endswith(SUPPORTED):
            return False
        return fnmatch.fnmatch(f, match) if match else True

    files = []
    if os.path.isdir(dataset_dir):
        files += [os.path.join(dataset_dir, f) for f in sorted(os.listdir(dataset_dir))
                  if keep(f)]
    if include_examples and os.path.isdir(EXAMPLE_DIR):
        files += [os.path.join(EXAMPLE_DIR, f) for f in sorted(os.listdir(EXAMPLE_DIR))
                  if keep(f)]
    return files


def main():
    ap = argparse.ArgumentParser(description="Run every dataset through the PVcopilot pipeline.")
    ap.add_argument("dataset_dir", nargs="?", default=DEFAULT_DATASET_DIR,
                    help="Folder of datasets to test (default: ./test_datasets/).")
    ap.add_argument("--include-examples", action="store_true",
                    help="Also test the bundled example datasets in ./data/.")
    ap.add_argument("--method", default=None,
                    choices=["YOY", "LR", "HW", "ARIMA", "CSD"],
                    help="Force a degradation method instead of the app's auto-gating.")
    ap.add_argument("--limit", type=int, default=None,
                    help="Only test the first N datasets (handy for a quick check).")
    ap.add_argument("--match", default=None,
                    help="Only test filenames matching this glob, e.g. 'my_*' or "
                         "'my_*_csv_12*.csv'. Combine with --limit for the first N matches.")
    args = ap.parse_args()

    files = collect_files(args.dataset_dir, args.include_examples, match=args.match)
    if args.limit:
        files = files[:args.limit]
    if not files:
        print(f"No datasets found in {args.dataset_dir} "
              f"(supported: {', '.join(SUPPORTED)}).")
        return 1

    # Run quietly: suppress the per-dataset chatter and the noisy prints from the
    # filter functions (they go to stdout), and show only a tqdm progress bar
    # (on stderr) with %done + ETA. The summary still prints at the end.
    results = []
    _bar = tqdm(files, desc="PVcopilot pipeline", unit="dataset", file=sys.stderr)
    with open(os.devnull, "w") as _devnull:
        for path in _bar:
            with redirect_stdout(_devnull):
                rec = run_one(path, forced_method=args.method)
            results.append(rec)

    counts = _counts(results)
    os.makedirs(REPORT_DIR, exist_ok=True)
    run = _next_run_number()    # 1, 2, 3, ... -- the highest number is the latest run
    meta = {
        "run": run,
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "dataset_dir": os.path.relpath(args.dataset_dir, PROJECT_ROOT),
        "n_files": len(files),
        "forced_method": args.method,
    }

    # Output files are numbered per run so the highest number is always the
    # latest. Results sheet = dataset | filtering result | degradation rate |
    # detected/created columns | why_missing.
    write_json(results, counts, meta,
               os.path.join(REPORT_DIR, f"pvcopilot_pipeline_report_{run}.json"))
    write_csv(results,
              os.path.join(REPORT_DIR, f"pvcopilot_pipeline_report_{run}.csv"))
    results_csv = os.path.join(REPORT_DIR, f"pvcopilot_results_{run}.csv")
    write_results_csv(results, results_csv)
    # Readable Excel copy: wrapped text, sized columns, frozen bold header,
    # amber tint on unreliable rows. Same data as the CSV.
    results_xlsx = os.path.join(REPORT_DIR, f"pvcopilot_results_{run}.xlsx")
    xlsx_ok = write_results_xlsx(results, results_xlsx)

    print("\nSummary: " + "   ".join(f"{k}: {counts[k]}" for k in STATUS_ORDER))
    print(f"Run #{run}. Results: {os.path.relpath(results_csv, PROJECT_ROOT)}"
          + (f" + {os.path.relpath(results_xlsx, PROJECT_ROOT)}" if xlsx_ok else "")
          + " (highest number = latest run).")
    # Non-zero exit if anything genuinely failed -- handy for CI / scripting.
    return 2 if counts["ERROR"] else 0


if __name__ == "__main__":
    sys.exit(main())
